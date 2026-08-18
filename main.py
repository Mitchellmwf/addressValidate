#streamlit run main.py

import streamlit as st
import re
import pycountry
import difflib
import zipfile
import openpyxl as op
import unicodedata
import datetime
import time
import json
from dateutil import parser as dateutil_parser
import uuid
from io import BytesIO

if st.session_state.get('startDate') is None:
    st.session_state.startDate = datetime.date.today()

if st.session_state.get('endDate') is None:
    st.session_state.endDate = datetime.date.today()

if st.session_state.get('streetTypeJSON') is None:
    st.session_state.customStreetTypeJSON = False
    with open("validTypes.json", "r") as f:
        st.session_state.streetTypeJSON = json.load(f)

if st.session_state.get('postalCodePatterns') is None:
    st.session_state.customPostalCodePatterns = False
    with open("postalCodes.json", "r") as f:
        st.session_state.postalCodePatterns = json.load(f)

if st.session_state.get('dateDoesntInvalidate') is None:
    st.session_state.dateDoesntInvalidate = False

def validateState(country, state):
    if not country:
        return False
    # Check if the country is valid
    try:
        countryObj = pycountry.countries.lookup(country)
    except LookupError:
        return False

    # Get subdivisions
    subdivisions = pycountry.subdivisions.get(country_code=countryObj.alpha_2)
    for subdivision in subdivisions:
        if subdivision.code.split('-')[1] == state:
            #return code if state is valid, otherwise return False
            return subdivision.code.split('-')[1]
    return False


def formatString(string, removeAccents=True, removePunctuation=True, toUpper=False, leaveApostrophes=False, removeSpaces=False):
    if not string:
        return None

    string = str(string).strip()

    if removeAccents:
        # Convert accented characters to their base characters
        string = unicodedata.normalize('NFKD', string)
        string = ''.join(
            c for c in string
            if not unicodedata.combining(c)
        )

    if removePunctuation:
        string = re.sub(r'[\\.,/()*?><;:]', '', string)
        string = re.sub(r'[`"]', "'", string)
        if not leaveApostrophes:
            string = re.sub(r"'", ' ', string)
        # Remove punctuation
        string = re.sub(r'[^\w\s]', ' ', string)

    # Replace multiple spaces with a single space
    string = re.sub(r'\s+', ' ', string)

    if removeSpaces:
        string = string.replace(" ", "")

    return string.upper() if toUpper else string

def validatePostalCode(country, postalCode):
    if not country:
        return False
    #format the postal code by removing accents and punctuation, and converting to uppercase, to improve regex matching
    postalCode = formatString(postalCode, toUpper=True)
    # Define regex patterns for postal codes based on country
    PSPatterns = st.session_state.postalCodePatterns
    
    pattern = PSPatterns.get(country)
    if not pattern:
        st.warning(f"No postal code pattern defined for country: {country}. Skipping postal code validation.")
        return True 
    matchCode = re.match(pattern, str(postalCode))
    matchCode2 = re.match(pattern, normalizePostalCode(str(postalCode), country))
    if pattern and matchCode:
        return matchCode.group(0)
    elif pattern and matchCode2:
        return matchCode2.group(0)
    return False

def normalizePostalCode(postalCode, country):
    if not postalCode or not country:
        return postalCode
    postalCode = formatString(postalCode, toUpper=True)
    # Use regex to insert the space in the correct position
    patterns = {
        'CA': r'^([A-Za-z]\d[A-Za-z])\s?(\d[A-Za-z]\d)$',  # Canada
        # Add more patterns for other countries that have formatting requirements that have strict formatting rules
        # UK won't work as it can e1g 4rte or sw1a 1aa
    }
    pattern = patterns.get(country)
    if not pattern:
        #change all Os to 0s and all Is to 1s
        pattern = r'^[A-Za-z0-9]+$'
    match = re.match(pattern, postalCode)
    if match:
        # If there are mutiple groups, join them with a space
        if len(match.groups()) > 1:
            return ' '.join(match.groups())
        else:
            return match.group(0)
    else:
        if country == 'CA':
            postalCode = formatString(postalCode, toUpper=True, removePunctuation=True)
            for i in range(len(postalCode)):
                if i == 1 or i == 4 or i == 6:
                    char = postalCode[i]
                    if char == 'O':
                        postalCode = postalCode.replace('O', '0')
                    elif char == 'I':
                        postalCode = postalCode.replace('I', '1')
                elif i == 0 or i == 2 or i == 5:
                    char = postalCode[i]
                    if char == '0':
                        postalCode = postalCode.replace('0', 'O')
                    elif char == '1':
                        postalCode = postalCode.replace('1', 'I')
            return postalCode
        else:
            return postalCode


def validateAddress(address):
    isExpired = None
    timestamp = None
    # If no address is provided, return an error message
    if not address:
        address['error'] = "No address provided."
        return address
    errorFields = []
    
    # If any of the required fields are missing, return an error message specifying which field is missing
    for field in ['country', 'state', 'city', 'postalCode', 'streetAddress', 'recordId', 'timestamp']:
        if address.get(field) == None or str(address.get(field)).strip().lower() == "none":
            errorFields.append(f"{field}")


    if errorFields:
        missingError = "Missing required fields: " + ", ".join(errorFields)
        address['error'] = missingError
        return address  


    try:
        # Give address as a dictionary with keys: country, state, postal_code, street_type
        try:
            country = pycountry.countries.lookup(formatString(address.get('country'), removeAccents=False, leaveApostrophes=True)) if address.get('country') else None
        except LookupError:
            try:
                countryCode = getCountryCode(address.get('country'))
                country = pycountry.countries.get(alpha_2=countryCode) if countryCode else None
            except LookupError:
                errorFields.append(f"Invalid country: {address.get('country')}")
            if not country:
                try:
                    countryNames = [country.name for country in pycountry.countries]
                    closestCountry = difflib.get_close_matches(formatString(address.get('country'), removeAccents=False, leaveApostrophes=True), countryNames, n=1, cutoff=0.8)
                    country = pycountry.countries.get(name=closestCountry[0]) if closestCountry else None
                except Exception as e:
                    errorFields.append(f"Invalid country: {address.get('country')}")
                
        countryCode = country.alpha_2 if country else None
        countryName = country.name.upper() if country else None
        if not country:
            errorFields.append(f"Invalid country: {address.get('country')}")
        state = address.get('state')
        if not state:
            errorFields.append("State/Province is required.")
        
        stateCode = getState(state, countryCode, "code") if address.get('state') else None
        stateName = getState(state, countryCode, "name") if address.get('state') else None

        city = address.get('city')
        if not city:
            errorFields.append("City is required.")
        postalCode = address.get('postalCode')
        if not postalCode:
            errorFields.append("Postal code is required.")
        streetAddress = address.get('streetAddress')
        if not streetAddress:
            errorFields.append("Street address is required.")
        
        
        streetValid = validateStreetType(abbreviateAddress(streetAddress))
        stateValid = validateState(countryCode, stateCode)
        postalCodeValid = validatePostalCode(countryCode, postalCode)

        if stateValid == False: 
            errorFields.append(f"Invalid state/province for the specified country {countryCode}/{stateCode}")

        if postalCodeValid == False:
            errorFields.append(f"Invalid postal code format for ({countryCode}): {postalCode}")
        
        if not streetValid:
            errorFields.append("Street type could not be determined from the address.")


        # Check if the timestamp is a valid date object
        if st.session_state.get('startDate') != st.session_state.get('endDate'): #ignore check if start and end dates are the same
            #check for null
            if not address.get('timestamp'):
                errorFields.append("Timestamp is required for date range validation.")

            #Check if its a datetime object, if not : try to parse
            elif address.get('timestamp') and not isinstance(address.get('timestamp'), (datetime.date, datetime.datetime)):
                parsedDate = parseDate(address.get('timestamp'))
                timestamp = parsedDate.date() if isinstance(parsedDate, datetime.datetime) else parsedDate
            elif address.get('timestamp') and isinstance(address.get('timestamp'), (datetime.date, datetime.datetime)):
                timestamp = address.get('timestamp')

            if not timestamp:
                errorFields.append(f"Timestamp {address.get('timestamp')} is not a valid date.")
            else:
                address['timestamp'] = timestamp

            # Check if the timestamp is within the start and end date range
            if address.get('timestamp') and isinstance(address.get('timestamp'), (datetime.date, datetime.datetime)):
                isExpired = not (st.session_state.startDate <= (address.get('timestamp').date() if isinstance(address.get('timestamp'), datetime.datetime) else address.get('timestamp')) <= st.session_state.endDate)
            else:
                isExpired = True

            #add as an error if the timestamp is not within the specified date range, unless the user has chosen to ignore date range validation
            if address.get('timestamp') and isExpired:
                #Skip if start and end dates are the same, as this is likely a single date range
                if not st.session_state.dateDoesntInvalidate:
                    errorFields.append(f"Timestamp {address.get('timestamp')} is not within the specified date range ({st.session_state.startDate} to {st.session_state.endDate}).")

        
        if errorFields:
            address['error'] = ", ".join(errorFields)
            return address

        # If all validations pass, return the validated address as a dictionary
        # but return state codes instead of names for US and Canada, as they are more commonly used in addresses
        if countryCode in ['US', 'CA']:
            stateName = stateCode
        validatedAddress = { 
            'country': countryName,
            'state': stateName.upper() if stateName else None,
            'postalCode': formatString(postalCodeValid, toUpper=True, removeAccents=True, leaveApostrophes=True),
            'streetAddress': formatString(streetValid, toUpper=True, removeAccents=True, leaveApostrophes=True),
            'city': formatString(city, toUpper=True, removeAccents=True, leaveApostrophes=True),
            'timestamp': address.get('timestamp'),
            'recordId': str(address.get('recordId')),
            'programId': str(address.get('programId')),
            'error': None,
            'expired': isExpired if st.session_state.startDate != st.session_state.endDate else False
        }
        ##st.success(f"{validatedAddress['streetAddress']}, {validatedAddress['city']} {validatedAddress['state']}, {validatedAddress['postalCode']}, {validatedAddress['country']} is valid.")
        return validatedAddress
    except Exception as e:
        address['error'] = f"An error occurred during validation: {str(e)}"
        return address

def parseDate(dateValue):
    if not dateValue:
        return None

    dateString = str(dateValue).strip()

    # Junk/placeholder values
    junk_values = {"n/a", "na", "none", "null", "-", "00/00/0000", "0000-00-00"}
    if dateString.lower() in junk_values:
        return None

    #Remove st, nd, rd, th from dates like 1st, 2nd, 3rd, 4th
    dateString = re.sub(r"(\d+)(st|nd|rd|th)", r"\1", dateString)

    # Explicit formats to try first (fast, unambiguous)
    known_formats = [
        "%Y-%m-%d",       # 2024-07-23
        "%m/%d/%Y",       # 07/23/2024
        "%d/%m/%Y",       # 23/07/2024
        "%Y/%m/%d",       # 2024/07/23
        "%B %d, %Y",      # July 23, 2024
        "%B %dst, %Y",    # March 5th, 2024 (handled separately below)
        "%b %d, %Y",      # Jul 23, 2024
        "%d-%m-%Y",       # 23-04-2023
        "%m-%d-%Y",       # 07-23-2024
        "%Y-%m-%d %H:%M:%S",
    ]

    for f in known_formats:
        try:
            return datetime.datetime.strptime(dateString, f)
        except ValueError:
            continue

    # dateutil's fuzzy parser as fallback
    # (e.g. "5-6-24", "2023/1/1", inconsistent spacing)
    try:
        return dateutil_parser.parse(dateString, fuzzy=True)
    except (ValueError, OverflowError):
        return None

    



def getCountryCode(countryName):
    if not countryName:
        return None
    
    countryName = str(countryName).strip()

    try:
        country = pycountry.countries.lookup(countryName)
        return country.alpha_2
    except LookupError:
        return None
    
def getState(stateName, countryCode, argumentType='name'):
    #Function that can take a state name or code (Allowing for minor misspellings) and a country code, and return the standardized state name or code based on the argumentType 
    if not stateName or not countryCode:
        return None
    try:
        # Avoid matching Quebec city, Mexico City, etc as states/provinces
        if ' city' in stateName.lower():
            return None
        stateName = formatString(str(stateName).strip())
        subdivisions = pycountry.subdivisions.get(country_code=countryCode)
        subdivisionNames = [subdivision.name for subdivision in subdivisions]
        subdivisionCodes = [subdivision.code.split('-')[1] for subdivision in subdivisions]
        #use difflib to find the closest match for the state name or code, allowing for minor misspellings
        closestName = difflib.get_close_matches(stateName, subdivisionNames, n=1, cutoff=0.8)
        closestCode = difflib.get_close_matches(stateName.upper(), subdivisionCodes, n=1, cutoff=0.8)
        if closestCode:
            closestSubdivision = next(subdivision for subdivision in subdivisions if subdivision.code.split('-')[1] == closestCode[0])
            return closestSubdivision.code.split('-')[1] if argumentType == 'code' else closestSubdivision.name
        elif closestName:
            closestSubdivision = next(subdivision for subdivision in subdivisions if subdivision.name == closestName[0])
            return closestSubdivision.code.split('-')[1] if argumentType == 'code' else closestSubdivision.name
        else:
            for sn in subdivisionNames:
                if stateName.lower() in sn.lower():
                    closestSubdivision = next(subdivision for subdivision in subdivisions if subdivision.name == sn)
                    return closestSubdivision.code.split('-')[1] if argumentType == 'code' else closestSubdivision.name
                elif formatString(stateName.lower()) in formatString(sn.lower()):
                    closestSubdivision = next(subdivision for subdivision in subdivisions if subdivision.name == sn)
                    return closestSubdivision.code.split('-')[1] if argumentType == 'code' else closestSubdivision.name
            return None
    except Exception as e:
        st.warning(f"An error occurred while validating state/province: {str(e)}")
        return None

def abbreviateAddress(address):
    # shorten things like south, north, east, west to S, N, E, W
    if not address:
        return None
    dictionary = {
        "south": "S",
        "north": "N",
        "east": "E",
        "west": "W",
        "northwest": "NW",
        "northeast": "NE",
        "southwest": "SW",
        "southeast": "SE",
        "s": "S",
        "n": "N",
        "e": "E",
        "w": "W",
        "nw": "NW",
        "ne": "NE",
        "sw": "SW",
        "se": "SE",
        "appartment": "APT",
        "app": "APT",
        "suite": "STE",
        "ste": "STE",
        "unit": "UNIT",
    }
    tokens = formatString(address).lower().replace(".", "").split()
    for token in tokens:
        if token in dictionary:
            address = re.sub(r'\b' + re.escape(token) + r'\b', dictionary[token], address, flags=re.IGNORECASE)
    return address.upper()

@st.cache_data
def validateStreetType(address):
    if not address:
        return False
    
    validTypes = st.session_state.streetTypeJSON

    # Split into tokens and check for street types, starting from the end of the string
    tokens = formatString(address).lower().replace(".", "").split()

    for token in reversed(tokens):
        if token in validTypes:
            #replace the street type in the address string with the standardized version
            standardizedType = validTypes[token]
            #remove the original street type from the address and replace with standardized version
            streetAddress = re.sub(r'\b' + re.escape(token) + r'\b', standardizedType, address, flags=re.IGNORECASE)
            return streetAddress.upper()

    #st.warning(formatString(address) + " | does not contain a valid street type.")
    return False

def fixSwappedCols(addressLineCol, cityCol, stateCol, postalCodeCol, countryRowCol, addressID):
    ##st.info(f"Attempting to fix swapped columns for the following data: {addressLineCol}, {cityCol}, {stateCol}, {postalCodeCol}, {countryRowCol}")
    foundCountry = None
    foundState = None
    foundPostalCode = None
    foundAddressLine = None
    foundCity = None
    foundDate = None
    errors = []
    unclaimed = []
    
    colList = [addressLineCol, cityCol, stateCol, postalCodeCol, countryRowCol]
    for i in range(len(colList)):
        currentCol = colList[i]
        testCountry = getCountryCode(currentCol)
        if testCountry:
            foundCountry = testCountry
            colList.pop(i)
            break
            
    for j in range(len(colList)):
        currentCol = formatString(str(colList[j]).strip(), removeAccents=True, removePunctuation=False)
        unclaimed.append(currentCol)
        testState = getState(currentCol, foundCountry, "name")
        #st.info(f"Testing {currentCol} for state with country {foundCountry}. Result: {testState}")
        testStreet = validateStreetType(abbreviateAddress(currentCol))

        #If column is a date or datetime, mark it as a date and remove from unclaimed list
        # check if the column is a datetime object
        if isinstance(colList[j], datetime.date):
            foundDate = colList[j]
            try:
                unclaimed.remove(currentCol)
            except ValueError:
                pass
        # If the column is a string that can be parsed as a date, mark it as a date and remove from unclaimed list
        if isinstance(colList[j], str):
            try:
                parsedDate = parseDate(colList[j])
                if parsedDate:
                    foundDate = parsedDate.date() if isinstance(parsedDate, datetime.datetime) else parsedDate
                    try:
                        unclaimed.remove(currentCol)
                    except ValueError:
                        pass
            except Exception as e:
                pass


        if testState:
            foundState = testState
            #if foundCity is the same as the state, the city match below likely miscaptured it
            if foundCity and foundCity.lower() == foundState.lower():
                    foundCity = None
                    errors.append("city")
            try:
                unclaimed.remove(currentCol)
            except ValueError:                
                pass

            
        if testStreet:
            try:
                errors.remove("address line")
            except ValueError:
                pass
            foundAddressLine = testStreet
            try:
                unclaimed.remove(currentCol)
            except ValueError:                
                pass

        #Check if 9 digits or shorter and contains digits, it's likely a postal code
        if len(currentCol) <= 9 and any(char.isdigit() for char in currentCol):
            foundPostalCode = normalizePostalCode(currentCol, foundCountry)
            try:
                unclaimed.remove(currentCol)
            except ValueError:                
                pass
        
        # if it contains no digits and is shorter than 12 characters, it's likely a city name
        if len(currentCol) < 12 and not any(char.isdigit() for char in currentCol) and foundState and foundCountry and currentCol.lower() != foundState.lower() and currentCol.lower() != foundCountry.lower():
            foundCity = currentCol
            try:
                unclaimed.remove(currentCol)
            except ValueError:                
                pass

        #If it contains spaces and digits and is longer than 8 characters, it's likely an address line
        if len(currentCol) > 8 and any(char.isdigit() for char in currentCol) and " " in currentCol and not testStreet:
            foundAddressLine = currentCol
            errors.append("address line")
            try:
                unclaimed.remove(currentCol)
            except ValueError:                
                pass

        if foundCountry and foundState and foundPostalCode and foundAddressLine and foundDate and len(unclaimed) == 1:
            foundCity = unclaimed[0]
            
    if foundCountry and foundState and foundPostalCode and foundAddressLine and foundCity and foundDate:
        # Check if the timestamp is within the start and end date range
        isExpired = not (st.session_state.startDate <= address.get('timestamp') <= st.session_state.endDate) if address.get('timestamp') else True
        address = {
            'country': foundCountry,
            'state': foundState,
            'postalCode': foundPostalCode,
            'streetAddress': foundAddressLine,
            'city': foundCity,
            'timestamp': foundDate,
            'error': errors.insert(0, "passed column swap fix") if errors else None,
            'recordId': str(addressID),
            'programId': f'{addressID}_{uuid.uuid4()}',
            'expired': isExpired
        }
        return address
    else:
        #st.warning("Could not automatically fix swapped columns. Please ensure the spreadsheet is formatted correctly.")
        #st.info(f"After attempting to fix swapped columns, the following data was found: Address line:{foundAddressLine}, City: {foundCity}, State: {foundState}, Postal code: {foundPostalCode}, Country: {foundCountry}")
        return None
    
def displayResults(validList, invalidList): 
    try:
    
        editMode = st.checkbox("Show edit mode", value=False)

        col1, col2 = st.columns(2)
        with col1:
            st.header("Valid addresses")
            valCon = st.container(height=600)
            for valid in validList:
                for col in ['streetAddress', 'city', 'state', 'country']:
                    if valid.get(col):
                        # Skip for state codes that are 2 letters, as they should be uppercase
                        if (col == 'state' and len(valid['state']) == 2) or col == 'postalCode':
                            continue
                        # Remove [] and () and their contents from the string, and convert to title case
                        valid[col] = re.sub(r'\[.*?\]|\(.*?\)', '', str(valid[col])).title()
                
                address = valid.copy()
                
                if editMode:
                    try:
                        with valCon.expander(f"{valid['recordId']} : {valid['streetAddress']}, {valid['city']} {valid['state']}, {valid['postalCode']}, {valid['country']}"):
                            for col in ['recordId', 'streetAddress', 'city', 'state', 'postalCode', 'country']:
                                if valid.get(col):
                                    address[col] = st.text_input(f"Edit {col}", value=str(address[col]), key=f"valid_{col}_{valid['programId']}")
                                else:
                                    address[col] = st.text_input(f"Edit {col}", value="", key=f"valid_{col}_{address['programId']}")
                            if valid.get('timestamp') and isinstance(valid.get('timestamp'), (datetime.date, datetime.datetime)):
                                if isinstance(valid.get('timestamp'), datetime.datetime):
                                    valid['timestamp'] = valid.get('timestamp').date()
                                address['timestamp'] = st.date_input("Edit timestamp", value=valid.get('timestamp') if valid.get('timestamp') else datetime.date.today(), key=f"valid_timestamp_{valid['programId']}", min_value=datetime.date(1900, 1, 1))
                            else:
                                st.write(f"Timestamp could not be interpreted: {address['timestamp']}. Please enter a valid date.")
                                address['timestamp'] = st.date_input("Edit timestamp", value=datetime.date.today(), key=f"valid_timestamp_{valid['programId']}", min_value=datetime.date(1900, 1, 1))                
                            #expired status checkbox
                            address['expired'] = st.checkbox("Expired", value=valid.get('expired', True), key=f"valid_expired_{valid['programId']}")
                            
                            # Save edits button
                            if st.button("Save edits", key=f"saveVal_{valid['programId']}"):
                                index = st.session_state.validList.index(valid)
                                st.session_state.validList[index] = address
                                st.success("Edits saved. You can test the changes or mark as valid if you think the address is now valid.")

                            #Test changes button
                            if st.button("Test changes", key=f"test_valid_{valid['programId']}"):
                                testResult = validateAddress(address)
                                if not testResult['error']:
                                    st.success(f"After changes, address is still valid: {testResult['streetAddress']}, {testResult['city']} {testResult['state']}, {testResult['postalCode']}, {testResult['country']}.")
                                    index = st.session_state.validList.index(valid)
                                    st.session_state.validList[index] = testResult
                                else:
                                    st.error(f"After changes, address is now invalid: {testResult['error']}. (Not saving changes)")

                            # Manually marking as invalid
                            if st.button("Mark as invalid", key=f"invalid_valid_{valid['programId']}"):
                                st.session_state.invalidList.append(address)
                                st.session_state.validList.remove(valid)
                                st.success(f"Address {valid['recordId']} marked as invalid. Moving...")
                                time.sleep(15)
                                refreshPage()
                    except Exception as e:
                        valCon.error(f"An error occurred while displaying a valid address: {str(e)}")
                else:
                    valCon.success(f"{valid['recordId']} : {valid['streetAddress']}, {valid['city']} {valid['state']}, {valid['postalCode']}, {valid['country']} is valid.")
                    # st.markdown(f"**Record ID: {valid['recordId']}**\n"
                    #     f"- Address: {valid['streetAddress']}, {valid['city']} {valid['state']}\n"
                    #     f"- Error: {valid.get('error_message', 'Unknown error')}")
                    
        with col2:
            st.header("Invalid addresses")

            invCon = st.container(height=600)
            for invalid in invalidList:
                address = invalid.copy()
                if editMode:
                    with invCon.expander(f"{invalid['recordId']} : {invalid['streetAddress']}, {invalid['city']} {invalid['state']}, {invalid['postalCode']}, {invalid['country']}"):
                        if invalid.get('error'):
                            st.warning(str(invalid['error']))

                        # Let user edit the fields
                        for col in ['recordId', 'streetAddress', 'city', 'state', 'postalCode', 'country']:
                            if address.get(col):
                                address[col] = st.text_input(f"Edit {col}", value=str(address[col]), key=f"{col}_{address['programId']}")
                            else:
                                address[col] = st.text_input(f"Edit {col}", value="", key=f"{col}_{address['programId']}")
                        if invalid.get('timestamp') and isinstance(invalid.get('timestamp'), datetime.date):
                            address['timestamp'] = st.date_input("Edit timestamp", value=invalid.get('timestamp') if invalid.get('timestamp') else datetime.date.today(), key=f"invalid_timestamp_{invalid['programId']}", min_value=datetime.date(1900, 1, 1))
                        else:
                            st.write(f"Timestamp could not be interpreted: {address['timestamp']}. Please enter a valid date.")
                            address['timestamp'] = st.date_input("Edit timestamp", value=datetime.date.today(), key=f"invalid_timestamp_{invalid['programId']}", min_value=datetime.date(1900, 1, 1))

                        #expired status checkbox
                        address['expired'] = st.checkbox("Expired", value=invalid.get('expired', True), key=f"invalid_expired_{invalid['programId']}")


                        # Save edits button
                        if st.button("Save edits", key=f"saveInv_{invalid['programId']}"):
                            index = st.session_state.invalidList.index(invalid)
                            st.session_state.invalidList[index] = address
                            st.success("Edits saved. You can test the changes or mark as valid if you think the address is now valid.")

                
                        # Button to test
                        if st.button("Test changes", key=f"test_{invalid['programId']}"):
                            testResult = validateAddress(address)
                            if not testResult['error']:
                                st.success(f"After changes, address is now valid: {testResult['streetAddress']}, {testResult['city']} {testResult['state']}, {testResult['postalCode']}, {testResult['country']}. Moving to valid list...")
                                st.session_state.validList.append(testResult)
                                st.session_state.invalidList.remove(invalid)
                                
                                time.sleep(2)
                                refreshPage()
                            else:
                                st.error(f"After changes, address is still invalid: {testResult['error']} (Changes not saved)")

                        # Manually marking as valid
                        if st.button("Mark as valid", key=f"valid_{address['programId']}"):
                            for col in ['streetAddress', 'city', 'state', 'country']:
                                if address.get(col):
                                    # Skip for state codes that are 2 letters, as they should be uppercase
                                    if (col == 'state' and len(address['state']) == 2) or col == 'postalCode':
                                        continue
                                    # Remove [] and () and their contents from the string, and convert to title case
                                    address[col] = re.sub(r'\[.*?\]|\(.*?\)', '', str(address[col])).title()
                            st.session_state.validList.append(address)
                            st.session_state.invalidList.remove(invalid)
                            st.success(f"Address {invalid['recordId']} marked as valid. Moving...")
                            time.sleep(2)
                            refreshPage()
                else:
                    invCon.error(f"{invalid['recordId']} : {invalid['streetAddress']}, {invalid['city']} {invalid['state']}, {invalid['postalCode']}, {invalid['country']} is invalid.")
                    # st.markdown(f"**Record ID: {invalid['recordId']}**\n"
                    #     f"- Address: {invalid['streetAddress']}, {invalid['city']} {invalid['state']}\n"
                    #     f"- Error: {invalid.get('error_message', 'Unknown error')}")
                    
    except Exception as e:
        st.error(f"An error occurred while displaying results: {str(e)}")


def saveResults(validList, invalidList):
    #Open file and load main sheet, set column widths, and add header row
    validFile = op.Workbook()
    sheet = validFile.active
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 22
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 6
    sheet.column_dimensions['I'].width = 50
    sheet.append(["Record ID", "Street Address", "City", "State/Province", "Postal Code", "Country", "Timestamp", "Expired"])

    # Format and save to sheet
    for i in range(len(validList)):
        valid = validList[i]

        # Make title case and format strings
        for col in ['streetAddress', 'city', 'state', 'country']:
            if valid.get(col):
                #Skip for state codes that are 2 letters, as they should be uppercase and postal Codes
                if (col == 'state' and len(valid['state']) == 2) or col == 'postalCode':
                    continue
                #remove [] and () and their contents from the string
                valid[col] = re.sub(r'\[.*?\]|\(.*?\)', '', valid[col]).title()

        #Convert recordId to int if possible, otherwise keep as string
        try:
            valid['recordId'] = int(valid['recordId'])
        except ValueError:
            valid['recordId'] = str(valid['recordId'])

        sheet.append([valid['recordId'], valid['streetAddress'], valid['city'], valid['state'] if len(str(valid['state'])) == 2 else valid['state'], valid['postalCode'], valid['country'], valid['timestamp'], "NO" if 'expired' in valid and not valid['expired'] else "YES"])

        #If the timestamp is a datetime object, set column G of current row to excel date format, otherwise leave as string
        if isinstance(valid['timestamp'], (datetime.datetime, datetime.date)):
            cell = sheet.cell(row=i+2, column=7)
            cell.number_format = 'YYYY-MM-DD'

    # Save the workbook to a file
    validBuffer = BytesIO()
    validFile.save(validBuffer)
    validBuffer.seek(0)

    #Open file and load main sheet, set column widths, and add header row
    invalidFile = op.Workbook()
    sheet = invalidFile.active
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 22
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 6
    sheet.column_dimensions['I'].width = 50
    sheet.append(["Record ID", "Street Address", "City", "State/Province", "Postal Code", "Country", "Timestamp", "Expired", "Error"])

    # Add each invalid address to the sheet without formatting, to preserve the original data for review
    for i in range(len(invalidList)):
        invalid = invalidList[i]

        #Convert recordId to int if possible, otherwise keep as string
        try:
            invalid['recordId'] = int(invalid['recordId'])
        except ValueError:
            invalid['recordId'] = str(invalid['recordId'])


        sheet.append([invalid['recordId'], invalid['streetAddress'], invalid['city'], invalid['state'], invalid['postalCode'], invalid['country'], invalid['timestamp'], "NO" if 'expired' in invalid and not invalid['expired'] else "YES", invalid['error'] if 'error' in invalid else None])
        #If the timestamp is a datetime object, set column G of current row to excel date format, otherwise leave as string
        if isinstance(invalid['timestamp'], (datetime.datetime, datetime.date)):
            cell = sheet.cell(row=i+2, column=7)
            cell.number_format = 'YYYY-MM-DD'

    # Save
    invalidBuffer = BytesIO()
    invalidFile.save(invalidBuffer)
    invalidBuffer.seek(0)

    zipBuffer = BytesIO()
    with zipfile.ZipFile(zipBuffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("valid_addresses.xlsx", validBuffer.getvalue())
        zf.writestr("invalid_addresses.xlsx", invalidBuffer.getvalue())
    zipBuffer.seek(0)

    return zipBuffer.getvalue()

    
    
def refreshPage():
    for key in list(st.session_state.keys()):
        if key.startswith(("address_", "city_", "state_", "postal_", "country_", "overwrite_", "test_", "valid_", "invalid_", "recordId_", "save_")):
            del st.session_state[key]
    st.rerun()
    ()
    
def mainPage():
    # Streamlit UI
    st.set_page_config(layout="centered")
    st.title("Address Validator")
    
    if st.session_state.customStreetTypeJSON:
        st.success("Using custom street type JSON file for validation.")
    if st.session_state.customPostalCodePatterns:
        st.success("Using custom postal code patterns JSON file for validation.")

    uploadedFile = st.file_uploader("Upload a spreadsheet with addresses", type=["xlsx"])

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        instructionBtn = st.button("Instructions")
    if instructionBtn:
        st.warning("Upload an Excel spreadsheet (.xlsx) with the following columns: Record ID, Address line 1, City, Province/State, Postal Code, Country. The first row should contain the column headers.")
        st.info("Click upload and select an Excel file with an xlsx extension.")
        st.info("After uploading the file, click 'Validate' to process the addresses. ")
        st.info("Once validation is complete, the page will refresh,\nThen you can choose to view or save the results. ")


    if st.session_state.get('validList') and st.session_state.get('invalidList'):
        with col4:
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            saveBtn = st.download_button("Download results as Excel files", data=saveResults(st.session_state.validList, st.session_state.invalidList), file_name=f"validated_addresses_{timestamp}.zip", mime="application/zip")
        if saveBtn:
            saveResults(st.session_state.validList, st.session_state.invalidList)
            st.success("Results saved as valid_addresses.xlsx and invalid_addresses.xlsx")

        with col3:
            displayBtn = st.button("Display results")
        if displayBtn:
            st.switch_page(reviewPage)

    with col2:
        validateBtn = st.button("Validate") and uploadedFile
    if validateBtn:
        try:
            # Process the uploaded file and validate addresses
            with st.progress(0, text="Processing"):
                progressBar = st.progress(0)
                # Read the Excel file using pandas
                # Excel rows: record id, Address line 1, City, Province/State, Postal Code, Country, timestamp
                sheet = op.load_workbook(uploadedFile).active
                validList = []
                invalidList = []
                # Iterate through the rows of the spreadsheet and validate each address
                currentRow = 0
                for row in sheet.iter_rows(values_only=True, min_row=2): #min_row=2 to skip header
                    rowCount = sheet.max_row - 1
                    currentRow += 1
                    progressBar.progress(currentRow / rowCount, text=f"Processing row {currentRow} of {rowCount}")
                    #skip empty rows
                    if row is None or all(cell is None for cell in row):
                        continue
                    # Grab columns from the row
                    addressLine = (row[1])
                    state = (row[3])
                    city = (row[2])
                    postalCode = ((row[4]))
                    country = row[5]
                    addressId = row[0]
                    timestamp = row[6]

                    #Assign columns to dict
                    address = {
                        'country': country,
                        'state': state,
                        'city': city,
                        'postalCode': postalCode,
                        'streetAddress': addressLine,
                        'recordId': addressId,
                        'timestamp': timestamp,
                        'error': None,
                        'programId': f'{addressId}_{uuid.uuid4()}' #Key to prevent duplicate keys in session state when editing addresses
                    }
                    validateResult = validateAddress(address)
                    if not validateResult['error']:
                        validList.append(validateResult)
                        ##st.write(f"{streetAddress} {streetType}, {state}, {postalCode}, {country} is valid.")
                    else:
                        tryFix = None
                        if  "missing" in validateResult['error']:
                            invalidList.append(address)
                            continue
                        elif isinstance(validateResult, list): # and len(validateResult) == 1:
                            tryFix = fixSwappedCols(addressLine, city, state, postalCode, country, addressId)
                        if tryFix:
                            newResult = validateAddress(tryFix)
                            if "passed column swap fix" in newResult['error']:
                                validList.append(newResult)
                            else:
                                invalidList.append(newResult)
                        else:
                            invalidList.append(validateResult)
                
            dummyAddress = {
                'country': 'N/A',
                'state': 'N/A',
                'city': 'N/A',
                'postalCode': 'N/A',
                'streetAddress': 'No addresses in category',
                'recordId': 'N/A',
                'timestamp': datetime.datetime.now(),
                'error': "No addresses found.",
                'programId': f'N/A_{uuid.uuid4()}',
                'expired': False
            }
            if invalidList and not validList:
                validList.append(dummyAddress)
                st.warning("No valid addresses found. A dummy address has been added to the valid list for display purposes.")
            elif not invalidList and validList:
                st.warning("No invalid addresses found. A dummy address has been added to the invalid list for display purposes.")
                invalidList.append(dummyAddress)
            elif not invalidList and not validList:
                st.warning("No valid or invalid addresses found. A dummy address has been added to both lists for display purposes.")
                validList.append(dummyAddress)
                invalidList.append(dummyAddress)

            # Update session state with results
            st.session_state.validList = validList
            st.session_state.invalidList = invalidList

            # After processing all addresses, display results
            st.success(f"Validation complete! Valid addresses: {len(st.session_state.validList)}, Invalid addresses: {len(st.session_state.invalidList)}")
            
            time.sleep(2)
            refreshPage()

        except Exception as e:
            st.error(f"An error occurred while processing the file: {str(e)}")
            if "tuple index" in str(e):
                st.warning("It seems there may be an issue with the column order or missing columns in the uploaded file. Please ensure the spreadsheet has the correct columns in the expected order.")
            st.info("Please ensure the uploaded file is a valid Excel spreadsheet with the correct columns.")

def reviewPage():
    if st.session_state.get('validList') and st.session_state.get('invalidList'):
        st.set_page_config(layout="wide")
    else:
        st.set_page_config(layout="centered")
    st.title("Review and Save Results")
    if st.button("Instructions"):
        st.info("This page displays the results of the address validation. Valid addresses are shown on the left, and invalid addresses are shown on the right. ")
        st.info("In editing mode, you can edit addresses and retest them. If an address is valid after editing, it will move to the valid list. You can also manually mark an address as valid. To enter editing mode, check the 'Show edit mode' checkbox. ")
        st.warning("Please note that marking an address as valid will save it as-is, it will not automatically fill blanks or be converted to a standardized format. Saving edits to an address only stores the changes to the exported excel file, it does not change the orignal data")
        st.warning("Although it is possible to edit the id, it is not recommended as it may cause issues with the database. It is best to keep the record ID as-is.")
    if st.session_state.get('validList') and st.session_state.get('invalidList'):
        displayResults(st.session_state.validList, st.session_state.invalidList)
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        if st.download_button("Download results as Excel files", data=saveResults(st.session_state.validList, st.session_state.invalidList), file_name=f"validated_addresses_{timestamp}.zip", mime="application/zip"):
            st.success("Results saved as valid_addresses.xlsx and invalid_addresses.xlsx")
        if st.button("Back to Validator"):
            st.switch_page(mainPage)
    else:
        st.info("No results to display. Please upload a file and validate addresses first.")
        if st.button("Go to Validator"):
            st.switch_page(mainPage)

def editParamsPage():
    st.title("Edit Parameters")
    st.info("This page allows you to edit the parameters used for address validation. You can add or remove valid street types. Please be careful when editing these parameters, as they affect the validation results. Date check is skipped when dates are identical.")

    st.subheader("Date range for address validation")
    st.session_state.startDate = st.date_input("Start date", value=st.session_state.startDate, min_value=datetime.date(1900, 1, 1))
    st.session_state.endDate = st.date_input("End date", value=st.session_state.endDate, min_value=st.session_state.startDate)
    if st.session_state.startDate > st.session_state.endDate:
        st.error("Start date cannot be after end date. Please select a valid date range.")
    if st.session_state.startDate != st.session_state.endDate or st.session_state.dateDoesntInvalidate:
        st.session_state.dateDoesntInvalidate = st.checkbox("Ignore dates for file output", value=st.session_state.dateDoesntInvalidate) 

    if st.session_state.startDate == st.session_state.endDate:
        st.warning("Start date and end date are the same. Date check will be skipped.")

    st.subheader("Valid Street Types")
    validTypes = st.session_state.streetTypeJSON

    groups = {}
    uploadedJSON = st.file_uploader("Upload a JSON file with street types", type=["json"])
    if uploadedJSON and st.button("Upload new street types JSON"):
        if uploadedJSON:
            try:
                newStreetTypes = json.load(uploadedJSON)
                if isinstance(newStreetTypes, dict):
                    st.session_state.streetTypeJSON = newStreetTypes
                    st.session_state.customStreetTypeJSON = True
                    st.success("Successfully updated street types from uploaded JSON.")
                    time.sleep(2)
                    refreshPage()
                else:
                    st.error("Uploaded JSON is not a valid dictionary of street types.")
            except Exception as e:
                st.error(f"An error occurred while loading the JSON: {str(e)}")

    # Display the valid street types in a text area for editing
    for abbreviation, standardized in validTypes.items():
        if standardized not in groups:
            groups[standardized] = []
        groups[standardized].append(abbreviation)

    st.header("Groups")
    for standardized, abbreviations in sorted(groups.items(), key=lambda x: x[0]):
        with st.expander(f"{standardized} ({', '.join(abbreviations)})"):
            standardized = st.text_input("Standardized street type", value=standardized, key=f"std_{standardized}")
            abbreviations = st.text_area("Abbreviations (one per line)", value="\n".join(abbreviations), key=f"abbr_{standardized}")
            if st.button("Save changes", key=f"save_{standardized}"):
                if standardized and abbreviations:
                    # Remove old abbreviations for this standardized type
                    for abbr in list(validTypes.keys()):
                        if validTypes[abbr] == standardized:
                            del validTypes[abbr]
                    # Add new abbreviations
                    for abbr in abbreviations.splitlines():
                        validTypes[abbr.strip()] = standardized.strip()
                    st.session_state.streetTypeJSON = validTypes
                    st.session_state.customStreetTypeJSON = True
                    st.success(f"Updated street type: {standardized} with abbreviations: {', '.join(abbreviations.splitlines())}")
                    time.sleep(2)
                    refreshPage()
                else:
                    st.error("Please enter both a standardized street type and at least one abbreviation.")
    
    with st.expander("Add a new street type"):
        newStandardized = st.text_input("New standardized street type", value="", key="new_std")
        newAbbreviations = st.text_area("New abbreviations (one per line)", value="", key="new_abbr")
        if st.button("Add new street type"):
            if newStandardized and newAbbreviations:
                for abbr in newAbbreviations.splitlines():
                    validTypes[abbr.strip()] = newStandardized.strip()
                st.success(f"Added new street type: {newStandardized} with abbreviations: {', '.join(newAbbreviations.splitlines())}")
                st.session_state.customStreetTypeJSON = True
                st.session_state.streetTypeJSON = validTypes
                time.sleep(2)
                refreshPage()
            else:
                st.error("Please enter both a standardized street type and at least one abbreviation.")


    if st.button("Go to Validator"):
        st.switch_page(mainPage)

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    settingsFile = f"street_types_{timestamp}.json"
    
    settingsBuffer = BytesIO()
    
    if st.download_button("Download settings", data=json.dumps(validTypes, indent=4), file_name=settingsFile):
        settingsBuffer.write(json.dumps(validTypes, indent=4).encode())
        settingsBuffer.seek(0)

def editPostalCodePage():

    uploadedJSON = st.file_uploader("Upload a JSON file with Postal Code patterns", type=["json"])
    if uploadedJSON and st.button("Upload new postal code patterns JSON"):
        #check if the uploaded file is a valid JSON
        try:
            newPostalCodePatterns = json.load(uploadedJSON)
            if isinstance(newPostalCodePatterns, dict):
                st.session_state.postalCodePatterns = newPostalCodePatterns
                st.session_state.customPostalCodePatterns = True
                st.success("Uploaded new postal code patterns JSON successfully.")
                time.sleep(2)
                refreshPage()
            else:
                st.error("Uploaded JSON is not a valid dictionary of postal code patterns.")
        except json.JSONDecodeError:
            st.error("Uploaded file is not a valid JSON.")

    st.title("Edit Postal Code Parameters")
    st.info("This page allows you to edit the parameters used for postal code validation. You can add or remove valid postal code formats. Please be careful when editing these parameters, as they may affect the validation results.")

    for country, pattern in st.session_state.postalCodePatterns.items():
        countryName = pycountry.countries.get(alpha_2=country).name if pycountry.countries.get(alpha_2=country) else country
        newPattern = st.text_input(f"Postal code pattern for {countryName}", value=pattern, key=f"postal_{country}")
        if newPattern != pattern:
            st.session_state.postalCodePatterns[country] = newPattern
            st.success(f"Updated postal code pattern for {countryName} to: {newPattern}")
            st.session_state.customPostalCodePatterns = True
            time.sleep(2)
            refreshPage()
    
    st.header("Add new postal code")
    addCountry = st.text_input("Country name or code")
    addPattern = st.text_input("Postal Code REGEX")
    # Check if already a country code
    if addCountry and len(addCountry) == 2 and pycountry.countries.get(alpha_2=addCountry.upper()):
        addCountry = addCountry.upper()
    else:
        countryCode = getCountryCode(addCountry)
        if countryCode:
            addCountry = countryCode
        else:
            st.warning(f"Could not find a valid country code for {addCountry}. Please enter a valid country name or ISO 3166-1 alpha-2 code.")

    if st.button("Add new parameter"):
        if addCountry and addPattern:
            st.session_state.postalCodePatterns[addCountry] = addPattern
            st.session_state.customPostalCodePatterns = True
            st.success(f"Added postal code pattern for {addCountry}: {addPattern}")
            time.sleep(2)
            refreshPage()

    if st.button("Go to Validator"):
        st.switch_page(mainPage)

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    settingsFile = f"PostalCodePatterns_{timestamp}.json"
    
    settingsBuffer = BytesIO()
    
    if st.download_button("Download settings", data=json.dumps(st.session_state.postalCodePatterns, indent=4), file_name=settingsFile):
        settingsBuffer.write(json.dumps(st.session_state.postalCodePatterns, indent=4).encode())
        settingsBuffer.seek(0)
    
def instructionsPage():
    # Renders the instructions.html file in the Streamlit app
    st.title("Instructions")

    st.header("Validation Instructions")

    st.markdown("**Step 1:** Click **Upload** to open your file browser.")
    st.image("instructions/images/step1.gif", use_container_width=True)

    st.markdown("""
    **Step 2:** Select the Excel file you want to validate and click **Open**.

    **Note:** The file must be in **.xlsx** format.
    """)

    st.image("instructions/images/step2.png", use_container_width=True)

    st.info(
        "Ensure that your Excel file contains the correct columns for address validation."
    )

    st.image(
        "instructions/images/columns.png",
        caption="Required Columns: ID, Address, City, State, Postal Code, and Country",
        use_container_width=True,
    )

    st.markdown("""
    **Step 3:** After selecting the file, click the **Validate** button to start the validation process.

    Once the page refreshes, you will have the option to view or export the results.
    """)

    st.image("instructions/images/step3.jpg", use_container_width=True)

    st.divider()

    st.header("Review and Save Results Instructions")

    st.markdown("""
    **Step 1:** To navigate to the review page, click the **Display Results** button after the validation process is complete, or use the **Review Results** link in the navigation menu.
    """)

    st.markdown("""
    **Step 2:** In the default view you will see:

    - ✅ Valid addresses on the left (green)
    - ❌ Invalid addresses on the right (red)

    Scroll through the tables to review the validation results.
    """)

    st.image(
        "instructions/images/review1.jpg",
        caption="Review Page",
        use_container_width=True,
    )

    st.markdown("""
    **Step 3:** Enable editing by checking the **Show edit mode** checkbox above the results. In edit mode you can modify address fields directly within the table.
    """)

    st.warning("Important Notes")

    st.markdown("""
    - Testing an **invalid address** automatically moves it to the valid table if the edits are valid.
    - Testing an **already valid address** does not automatically save or move it.
    - Marking an address as valid saves it **as entered**. Missing fields will **not** be automatically filled or standardized.
    - Saving edits only affects the exported Excel file—it does **not** modify the original data.
    - Editing the **ID** is possible but **not recommended**, as it may cause database issues.
    - For best results, use full address names (e.g., **Ontario** instead of **ON**, **Street** instead of **St.**).
    """)

    st.markdown("""
    **Step 4:** After reviewing and editing the addresses, click **Download Results** to export the validated data as zipped Excel files.
    """)

    st.divider()

    st.header("Street Type Parameters")

    st.write(
        "This page provides information about the street type parameters used during address validation."
    )

    st.write(
        'Street types are standardized abbreviations for common street suffixes, such as **"St"** for **"Street"** or **"Ave"** for **"Avenue"**.'
    )

    with st.expander("Uploading a File from Previous Edits", expanded=True):

        st.write(
            "If you previously exported your validation results, you can upload them again for additional edits or validation."
        )

        st.write(
            "The file must be in **.json** format and should have been exported from the address validation tool."
        )

        st.info(
            'For a visual guide, refer to the "Validation Instructions" section above.'
        )

    with st.expander("Making Edits", expanded=True):

        st.markdown("""
    **Step 1:** Find the standardized street type you want to edit (e.g. **St → Street**).

    If you want to add a new street type instead, see the section below.
    """)

        st.markdown("""
    **Step 2:** Click the table cell containing the street type. The cell will become editable.
    """)

        st.markdown("""
    **Step 3:** Each entry contains:

    - Abbreviation
    - Full street type name

    To add additional forms, separate each one with a **new line only**.

    **Note:** Download the updated settings after editing so they can be reused in future validation sessions.
    """)

        st.image(
            "instructions/images/streetTypeEdit.jpg",
            caption="Street Type Editing",
            use_container_width=True,
        )

    with st.expander("Adding New Street Types", expanded=True):

        st.markdown("""
    **Step 1:** Scroll to the bottom of the street type table to find the **Add New Street Type** section.
    """)

        st.markdown("""
    **Step 2:** Enter:

    - Abbreviation
    - Full street type name

    If multiple forms exist, separate each one using a new line.
    """)

        st.info(
            "Copy the standardized abbreviation into the abbreviation field and download the updated settings when finished."
        )

        st.image(
            "instructions/images/streetTypeAdd.jpg",
            caption="Adding Street Types",
            use_container_width=True,
        )

#Set pages and navigation
editPage = st.Page(editParamsPage, title="Edit Parameters", url_path="edit-params")
reviewPage = st.Page(reviewPage, title="Review Results", url_path="review")
mainPage = st.Page(mainPage, title="Validator", url_path="validator", default=True)
postalCodePage = st.Page(editPostalCodePage, title="Edit Postal Codes", url_path="edit-postal-codes")
site = st.navigation([
    mainPage,
    reviewPage,
    editPage,
    instructionsPage,
], expanded=True, position="top" 
)
#launch the site
site.run()